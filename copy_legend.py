import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy
from tqdm import tqdm


def copy_sheet_to_files(source_file, sheet_name, target_folder, target_sheet_name):
    # Carica il foglio Excel di origine
    source_workbook = load_workbook(source_file)
    source_sheet = source_workbook[sheet_name]

    # Elabora tutti i file Excel nella cartella target
    for filename in os.listdir(target_folder):

        if filename.endswith(".xlsx") and filename != source_file:
            target_file = os.path.join(target_folder, filename)
            target_workbook = load_workbook(target_file)
            if target_sheet_name not in target_workbook.sheetnames:
                target_workbook.create_sheet(target_sheet_name)

            city = filename[8:-5]
            print(city)
            punteggi = target_workbook['Punteggi']
            indirizzo = punteggi['C5'].value
            
            target_sheet = target_workbook[target_sheet_name]

            if punteggi['A1'].value != 'SCHEDA DI VALUTAZIONE':
                # Inserisci quattro righe sopra la prima riga
                punteggi.insert_rows(1, amount=4)
                punteggi['A1'].value = 'SCHEDA DI VALUTAZIONE'

            # Copia i dati dalla cella a quella mantenendo la formattazione
            for row in source_sheet.iter_rows():
                for cell in row:
                                      
                    target_cell = target_sheet[cell.coordinate]
                    target_cell.value = cell.value

                    # Copia la formattazione dalla cella di origine alla cella di destinazione
                    if cell.has_style:
                        target_cell.font = copy(cell.font)
                        target_cell.border = copy(cell.border)
                        target_cell.fill = copy(cell.fill)
                        target_cell.number_format = copy(cell.number_format)
                        target_cell.protection = copy(cell.protection)
                        target_cell.alignment = copy(cell.alignment)

            target_workbook.save(target_file)
            target_workbook.close()


if __name__ == "__main__":
    source_file = "mistery shop\\Mistery_ALESSANDRIA_REVmw.xlsx"
    source_sheet_name = "Risultati sintetici per negozio"
    target_directory = "mistery shop\\data"
    target_sheet_name = "Indici"

    copy_sheet_to_files(source_file, source_sheet_name,
                        target_directory, target_sheet_name)
