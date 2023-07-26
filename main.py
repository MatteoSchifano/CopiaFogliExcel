import os
from openpyxl import load_workbook
from copy import copy
from tqdm import tqdm


def copy_sheet_to_files(source_file, sheet_name, target_folder, target_sheet_name):

    source_workbook = load_workbook(source_file)
    source_sheet = source_workbook[sheet_name]

    for filename in tqdm(os.listdir(target_folder)):

        if filename.endswith(".xlsx") and filename != source_file:

            target_file = os.path.join(target_folder, filename)
            target_workbook = load_workbook(target_file)

            if target_sheet_name not in target_workbook.sheetnames:
                target_workbook.create_sheet(target_sheet_name)

            target_sheet = target_workbook[target_sheet_name]

            for row in source_sheet.iter_rows():
                for cell in row:

                    target_cell = target_sheet[cell.coordinate]
                    target_cell.value = cell.value

                    if cell.has_style:
                        target_cell.font = copy(cell.font)
                        target_cell.border = copy(cell.border)
                        target_cell.fill = copy(cell.fill)
                        target_cell.number_format = copy(cell.number_format)
                        target_cell.protection = copy(cell.protection)
                        target_cell.alignment = copy(cell.alignment)

            target_workbook.save(target_file)
            target_workbook.close()


if __name__ == "__main__":
    source_file = "source.xlsx"
    source_sheet_name = "sheet source"
    target_directory = "data"
    target_sheet_name = "target sheet"

    copy_sheet_to_files(source_file, source_sheet_name,
                        target_directory, target_sheet_name)
